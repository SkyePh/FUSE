from fastapi import FastAPI, Request, Form, BackgroundTasks, Query
from fastapi.responses import RedirectResponse, FileResponse
from contextlib import asynccontextmanager
from typing import List, Optional
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import asyncio
from fastapi.templating import Jinja2Templates
from starlette.middleware.sessions import SessionMiddleware
from fastapi.staticfiles import StaticFiles
from starlette.responses import HTMLResponse
from scraper import scrape_eu_portal
import sys
import os
from datetime import datetime, date
from urllib.parse import urlencode
import json
from database import fetch_all_calls, fetch_calls_by_filters, get_pool

# Windows-specific fix for Playwright subprocess execution
if sys.platform == "win32":
    asyncio.set_event_loop_policy(asyncio.WindowsSelectorEventLoopPolicy())

app = FastAPI()

app.add_middleware(SessionMiddleware, secret_key="secretkey")

# Set up Jinja2 templates
templates = Jinja2Templates(directory="templates")

def format_date(value):
    if value is None:
        return ""
    try:
        # If value is already a date or datetime object, use it directly.
        if isinstance(value, (date, datetime)):
            return value.strftime("%d %b %Y").upper()
        # Otherwise assume it's an ISO string in the format "YYYY-MM-DD"
        return datetime.strptime(value, "%Y-%m-%d").strftime("%d %b %Y").upper()
    except Exception as e:
        print("format_date error:", e)
        return value

templates.env.filters['format_date'] = format_date


# Serve static files (if needed)
app.mount("/static", StaticFiles(directory="static"), name="static")


@app.get("/")
async def root():
    """ Redirect root to home """
    return RedirectResponse(url="/home")

from fastapi.responses import RedirectResponse

@app.get("/home")
async def home(request: Request):

    return templates.TemplateResponse("home.html", {"request": request})

@app.get("/choose_closed")
async def choose_closed(request: Request):

    return templates.TemplateResponse("choose_closed.html", {"request": request})


@app.post("/fetch-categories")
async def fetch_categories(
    request: Request,
    closed: str = Form("false"),
    forthcoming: str = Form("false"),
    open_: str = Form("false", alias="open"),
    keyword: Optional[str] = Form("")
):
    """
    Fetch categories and pass them directly to /categories via query parameters.
    """

    print(f"Raw Data from Form Submission → Closed: {closed}, Forthcoming: {forthcoming}, Open: {open_}")

    # Convert "true"/"false" strings to actual booleans
    closed_bool = closed.lower() == "true"
    forthcoming_bool = forthcoming.lower() == "true"
    open_bool = open_.lower() == "true"

    print(f"Converted Bools → Closed: {closed_bool}, Forthcoming: {forthcoming_bool}, Open: {open_bool}")

    # Fetch categories from scraper
    categories = await scrape_eu_portal(
        get_categories_only=True,
        closed_option=closed_bool,
        forthcoming_option=forthcoming_bool,
        open_option=open_bool,
        keyword=keyword
    )

    print("Categories Retrieved from Scraper:", categories)

    return templates.TemplateResponse("options.html", {
        "request": request,
        "categories": categories,
        "closed": closed_bool,
        "forthcoming": forthcoming_bool,
        "open": open_bool
    })


@app.get("/categories")
async def categories_page(
    request: Request,
    categories: List[str] = Query([]),
    closed: bool = Query(False),
    forthcoming: bool = Query(False),
    open_: bool = Query(False, alias="open")
):
    """
    Load category selection page using query parameters.
    """
    print("Categories from Query Params:", categories)
    print(f"Filters - Closed: {closed}, Forthcoming: {forthcoming}, Open: {open_}")

    return templates.TemplateResponse("options.html", {
        "request": request,
        "categories": categories,
        "closed": closed,
        "forthcoming": forthcoming,
        "open": open_
    })

@app.get("/loading")
async def loading_page(request: Request, redirect_url: str):
    """
    Displays a loading screen and redirects to the specified URL after a delay.
    """

    progress_flag = "scraping_in_progress.json"

    #wait_interval = 3

    if not os.path.exists(progress_flag):
        return RedirectResponse(url=redirect_url)

    #await asyncio.sleep(wait_interval)

    return templates.TemplateResponse("loading.html", {"request": request, "redirect_url": redirect_url})


@app.post("/scrape")
async def scrape_endpoint(
    request: Request,
    background_tasks: BackgroundTasks,
    categories: List[str] = Form(...),
    closed: bool = Form(...),
    forthcoming: bool = Form(...),
    open_: bool = Form(..., alias="open")
):
    """
    Starts scraping for selected categories in the background with filtering options.
    """
    print("Scrape Endpoint Called")
    print(f"Selected Categories: {categories}")
    print(f"Filters - Closed: {closed}, Forthcoming: {forthcoming}, Open: {open_}")

    # Start the scraper with the selected filters
    background_tasks.add_task(
        scrape_eu_portal,
        closed_option=closed,
        forthcoming_option=forthcoming,
        open_option=open_,
        desired_category=categories
    )

    return RedirectResponse(url="/loading?redirect_url=/results", status_code=303)


@app.get("/results")
async def get_results(request: Request):
    """
    Fetches the scraped results and displays them in an HTML page.
    """
    default_status=["open for submission", "forthcoming"]
    data = await fetch_calls_by_filters(keyword="", status=default_status, probability="All")

    return templates.TemplateResponse("results.html", {
        "request": request,
        "data": data,
        "selected_status": default_status,
        "keyword": "",
        "selected_probability": "all"
    })

@app.get("/results_ssbi")
async def results_iframe(request: Request):

    return templates.TemplateResponse("results_iframe.html", {"request": request})


@app.get("/search", response_class=HTMLResponse)
async def search_calls(
    request: Request,
    keyword: str = "",
    status: List[str] = Query([]),
    probability: str = "all"
):
    data = await fetch_calls_by_filters(keyword=keyword, status=status, probability=probability)
    return templates.TemplateResponse("results.html", {
        "request": request,
        "data": data,
        "selected_status": status,  # List of selected statuses
        "keyword": keyword,
        "selected_probability": probability
    })

def extract_group_name(identifier):
    """Extracts the category group from an identifier"""
    parts = identifier.split('-')
    if parts[0] == "HORIZON" and len(parts) > 1:
        return parts[1]  # Returns 'CL5' from 'HORIZON-CL5-D4'
    elif len(parts) > 1:
        return parts[0]  # Returns 'JU' or another top-level group
    else:
        return identifier  # Fallback for unknown formats

@app.get("/category/{cat_name}", response_class=HTMLResponse)
async def view_category(request: Request, cat_name: str):
    pool = await get_pool()
    async with pool.acquire() as conn:
        rows = await conn.fetch(
            """
            SELECT sc.identifier
            FROM scraped_calls AS sc
            INNER JOIN categories AS c ON sc.category_id = c.id
            WHERE lower(c.name) = lower($1)
            """,
            cat_name
        )
    identifiers = [row["identifier"] for row in rows]
    return templates.TemplateResponse("category.html", {
        "request": request,
        "category": cat_name,
        "identifiers": identifiers
    })

@app.get("/export-excel")
async def export_excel(
    keyword: str = "",
    statuses: str = "",
    probability: str = "all"
):
    output_excel_path = "scraped_results.xlsx"

    pool = await get_pool()
    async with pool.acquire() as conn:
        query = """
            SELECT
                sc.identifier,
                sc.title,
                sc.action_type,
                sc.budget,
                sc.funding_per_project,
                sc.deadline_primary,
                sc.deadline_secondary,
                sc.accepted_projects,
                sc.probability_rate,
                sc.link,
                sc.opening_date,
                c.name AS category_name,
                sc.status
            FROM scraped_calls sc
            INNER JOIN categories c ON sc.category_id = c.id
            WHERE 1=1
        """
        params = []
        idx = 1

        if keyword.strip():
            query += (
                f" AND (lower(sc.identifier) ILIKE '%' || ${idx} || '%' "
                f"OR lower(sc.title) ILIKE '%' || ${idx} || '%')"
            )
            params.append(keyword)
            idx += 1

        if statuses:
            status_list = [s.strip().lower() for s in statuses.split(",") if s.strip()]
            query += f" AND lower(sc.status) = ANY(${idx})"
            params.append(status_list)
            idx += 1

        if probability.lower() != "all":
            query += f" AND lower(sc.probability_rate) = ${idx}"
            params.append(probability.lower())
            idx += 1

        rows = await conn.fetch(query, *params)
        data = [dict(r) for r in rows]

    if not data:
        return {"error": "No data found"}

    grouped = {}
    for entry in data:
        group = entry["category_name"] or "Unknown"
        grouped.setdefault(group, []).append(entry)

    with pd.ExcelWriter(output_excel_path) as writer:
        for cat, entries in grouped.items():
            df = pd.DataFrame(entries)
            df.drop(columns=["id"], errors="ignore", inplace=True)
            df.to_excel(writer, sheet_name=cat[:31], index=False)

    wb = load_workbook(output_excel_path)
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        headers = [c.value for c in ws[1]]
        if "Probability Rate" in headers:
            col = headers.index("Probability Rate") + 1
            for r in range(2, ws.max_row + 1):
                cell = ws.cell(row=r, column=col)
                color = {"low":"ADD8E6","medium":"FFFF00","high":"90EE90"}.get((cell.value or "").lower())
                if color:
                    cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
    wb.save(output_excel_path)

    return FileResponse(
        output_excel_path,
        filename="filtered_results.xlsx",
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

# Run with: uvicorn api:app --host 127.0.0.1 --port 5000 (DONT USE --reload)
